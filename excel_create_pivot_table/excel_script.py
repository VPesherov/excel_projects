import openpyxl
import win32com.client as win32
from openpyxl import Workbook
from openpyxl.styles import Font
import os

import functions


data = {
    "A": ["Наименование товара", 40],
    "B": ["Склад", 50],
    "C": ["Город", 30],
    "D": ["Поставщик", 40],
    "E": ["Сумма, р.", 15],
    "F": ["Количество", 15],
    "G": ["Цена за единицу", 17],
}


def insert_pivot_table_field_set(pt):
    field_rows = {
        data.get("A")[0]: pt.PivotFields(data.get("A")[0])
    }
    field_values = {
        data.get("G")[0]: pt.PivotFields(data.get("G")[0])
    }
    field_filters = {
        data.get("D")[0]: pt.PivotFields(data.get("D")[0]),
        data.get("B")[0]: pt.PivotFields(data.get("B")[0])
    }
    field_column = {
        data.get("C")[0]: pt.PivotFields(data.get("C")[0])
    }

    # 1 - Строки, 2 - Столбцы 3 - фильтры 4 - Значения
    field_values[data.get("G")[0]].Orientation = 4
    # XlConsolidationFunction
    field_values[data.get("G")[0]].Function = -4106
    field_values[data.get("G")[0]].NumberFormat = "_-* # ##0,00_-;-* # ##0,00_-;_-* \"-\"??_-;_-@_-"

    field_rows[data.get("A")[0]].Orientation = 1

    field_column[data.get("C")[0]].Orientation = 2
    field_filters[data.get("D")[0]].Orientation = 3
    field_filters[data.get("B")[0]].Orientation = 3


def clear_pts(ws):
    for pt in ws.PivotTables():
        pt.TableRange2.Clear()


def create_pivot_table(output_file_name, directory=None):

    xl_app = win32.Dispatch('Excel.Application')
    if directory:
        wb = xl_app.Workbooks.Open(directory + '/' + output_file_name)
    else:
        my_path = os.path.abspath(os.curdir)
        wb = xl_app.Workbooks.Open(my_path + '\\' + output_file_name)

    ws_data = wb.Worksheets("Данные")
    ws_report = wb.Worksheets("Сводная")
    pt_cache = wb.PivotCaches().Create(1, ws_data.Range("A1").CurrentRegion)
    pt = pt_cache.CreatePivotTable(ws_report.Range("A3"), "Сводная таблица сравнение")
    pt.RowAxisLayout(1)

    insert_pivot_table_field_set(pt)
    wb.Close(True)
    xl_app.Application.Quit()


def work_with_excel(my_path, output_file_name, directory=None):
    wb = openpyxl.reader.excel.load_workbook(filename=my_path, data_only=True)
    wb.active = 1
    sheet = wb.active
    colour = sheet['B6'].fill.start_color.index  # Green Color
    green_product = []
    for i in range(1, sheet.max_row + 1):
        if sheet['B' + str(i)].fill.start_color.index == colour:
            green_product.append(sheet['A' + str(i)].value)

    wb.active = 0
    sheet = wb.active

    wb_1 = Workbook()

    wb_1.create_sheet("Данные", 0)
    wb_1.create_sheet("Сводная", 1)
    wb_1.active = 0
    sheet_1 = wb_1.active
    for i in data:
        sheet_1[i + "1"] = data.get(i)[0]

    for column in range(7):
        sheet_1[1][column].font = Font(bold=True)

    i = 1
    for row in range(4, sheet.max_row + 1):
        product_name = sheet[row][0].value
        if 'всего' in str(product_name).lower():
            continue
        if product_name is not None:
            product_const = product_name
        if product_const not in green_product:
            continue
        if 'итого' in str(product_name).lower():
            continue
        storage = sheet[row][1].value
        if storage is not None:
            storage_const = storage
        if 'всего' in str(storage).lower():
            continue
        provider = sheet[row][2].value
        if provider == 'Излишки по инвентаризации':
            continue

        quantity = sheet[row][3].value
        amount = sheet[row][4].value
        cost_per_unit = quantity / amount

        i += 1
        sheet_1["A" + str(i)] = product_const
        sheet_1["B" + str(i)] = storage_const
        sheet_1["C" + str(i)] = functions.city_from_brackets(str(storage_const))
        sheet_1["D" + str(i)] = provider
        sheet_1["E" + str(i)] = quantity
        sheet_1["F" + str(i)] = amount
        sheet_1["G" + str(i)] = float(functions.to_fixed(cost_per_unit, 2))

    for i in data:
        sheet_1.column_dimensions[i].width = data.get(i)[1]

    if directory:
        wb_1.save(directory + '/' + output_file_name)
        workbook = openpyxl.load_workbook(directory + '/' + output_file_name)
        del workbook['Sheet']
        workbook.save(directory + '/' + output_file_name)
        create_pivot_table(output_file_name, directory)
    else:
        wb_1.save(output_file_name)
        workbook = openpyxl.load_workbook(output_file_name)
        del workbook['Sheet']
        workbook.save(output_file_name)
        create_pivot_table(output_file_name)
    # del workbook['Sheet']
    # if directory:
    #     workbook.save(output_file_name)
    # else:
    #     print(directory + '/' + output_file_name)
    #     workbook.save(directory + '/' + output_file_name)
